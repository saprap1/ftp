#Get data from pro football reference
import requests
import lxml.html as lh
import pandas as pd
import re
from bs4 import BeautifulSoup
import requests
import openpyxl
import os
import string

from datetime import datetime
from datetime import timedelta 


#FROM HONORS PROJECT
def getSheet(link_first, link_second, version, name, dist_feet, avg_speed, posts, drives, year, avg_min):
    link = link_first + version + link_second
    a = requests.get(link)
    soup = BeautifulSoup(a.text, 'lxml')

    tb = soup.find("tbody")
    try:
        row = tb.findAll("tr")
        getW = soup.find('span', {'itemprop': 'weight'})
        weight = getW.text[0:3]
    except:
        v = int(version) + 1
        version = "0" + str(v)
        #print (link + " " + version)
        getSheet(link_first, link_second, version, name, dist_feet, avg_speed, posts, drives, year, avg_min)
        print(name)
        return
    global counting
    counting += 1
    print(name + " " + str(counting))
    labels = ['GM', 'Date', 'Weight', 'MP', 'FGA', '3PA', 'FTA', 'ORB', 'DRB', 'AST', 'TO', 'Fouls', 'PTS', 'dist_feet', 'avg_speed', "post_ups", "drives", "Injury"]
    
    bigL = []
    for x in row:
        items = x.findAll("td")
        counter = 0
        L = []
        for y in items:
            if(counter == 0):
                if(y.string=="None"):
                    print(y.string)
                    continue

            '''
            0 = game
            1 = date
            8 = minutes played
            10 = fga
            13 = 3fga
            16 = fta
            18 = offensive rebounds
            19 = def rebounds
            21 = assists
            24 = turnovers
            25 = personal fouls
            26 = points
            '''

            stats_to_ints = [0, 10, 13, 16, 18, 19, 21, 24, 25, 26]
            date = 1
            minutes = 8
            
            
            if (counter in stats_to_ints or counter == date or counter == minutes):
                if (counter in stats_to_ints):
                    x = y.string
                    if (x!=None):
                        L.append(int(x))
                    else:
                        L.append("")
                else:
                    if(counter==minutes):
                        s = y.text.split(":")
                        sInt0 = float(s[0])
                        sInt1 = float(s[1])
                        sInt1 /= 60
                        sInt1 = round(sInt1, 2)
                        
                        sInt0 += sInt1
                        L.append(sInt0)
                        
                        divisor = sInt0/avg_min
                        divisor = (round(divisor, 2))
                        
                    #date appending
                    else:
                        L.append(y.text)
                        L.append(int(weight))
            counter+=1
        if (len(L)>3):
            #float(str(round(answer, 2)))
            
            L.append(float(dist_feet)*divisor)
            L.append(float(avg_speed))
            L.append(float(posts)*divisor)
            L.append(float(drives)*divisor)
        if(len(L)==3):
            zero = [0,0,0,0,0,0,0,0,0,0,0,0,0,0]
            L.extend(zero)

        if(len(L)!=0):
            L.append("")
            #L.append("")
            bigL.append(L)
            

    df = pd.DataFrame.from_records(bigL, columns=labels)
    df.dropna()
    #print(df)
    filename = name + ".xlsx"
    if(year=="2014"):
        path = "bball_reference_2013-2014/" + filename
    elif year=="2015":
        path = "bball_reference_2014-2015/" + filename
    elif year=="2019":
        path = "bball_reference_2018-2019/" + filename
        
    writer = pd.ExcelWriter(path)
    df.to_excel(writer,'Sheet1')
    writer.save()

# FROM HONORS PROJECT
def loadSportsVu(book, sheet, row_count, year):
    
    for i in range (2, row_count+1):
        s = sheet.cell(row=i, column=2).value
        if (s == None):
            continue
        #call function on s
        name = s[1:]
        
        name = re.sub(r'[^\w\s]','',name)
        
        if(name == "JJ Barea"):
            name = "Jose Barea"
        elif(name=="Henry Walker"):
            name = "Bill Walker"
        elif(name == "Luc Mbah a Moute"):
            name = "Luc Mbaha"
        elif(name == "Mo Williams"):
            name = "Maurice Williams"
        elif(name == "Nene"):
            name = "Nene Hilario"
        elif(name == "Cedi Osman"):
            name = "De-Cedi Osman"
        elif(name == "Clint Capela"):
            name = "Ca-Clint Capela"
        elif(name == "Frank Ntilikina"):
            name = "La-Frank Ntilikina"
        elif(name == "Maxi Kleber"):
            name = "Maxi Klebi"
            
        arr = name.split(" ")
        version = "01"
        #column 5 and 6
        avg_min = sheet.cell(row=i, column=4).value
        dist_feet = sheet.cell(row=i, column=5).value
        avg_speed = sheet.cell(row=i, column=6).value
        posts = sheet.cell(row=i, column=7).value
        drives = sheet.cell(row=i, column=8).value
        
        if (len(arr[1])>=5):
            link_first = ("https://www.basketball-reference.com/players/" + arr[0][0] + "/" + arr[1][0:5] + arr[0][0:2])
            #link_first += version
            link_second = "/gamelog/" + year + "/"
            link_first = link_first.lower()
            link_second = link_second.lower()
        else:
            link_first = ("https://www.basketball-reference.com/players/" + arr[0][0] + "/" + arr[1] + arr[0][0:2])
            link_second = "/gamelog/" + year + "/"
            link_first = link_first.lower()
            link_second = link_second.lower()
        f = name + "_" + year
        getSheet(link_first, link_second, version, f, dist_feet, avg_speed, posts, drives, year, avg_min)


'''
def loadPlayerWorkbook(name, link, wksht):
    a = requests.get(link)

    print(link)

    soup = BeautifulSoup(a.text, 'lxml')

    tb = soup.find("tbody")
    row = tb.findAll("tr")
    getW = soup.find('span', {'itemprop': 'weight'})
    weight = getW.text[0:3]
    print(weight)
    labels = ["Name", "Date", "G#", "Cmp", "Att", "Cmp%", "Yds", "TD", "Int", "Rate", "Sk", "Yds", "Y/A", "AY/A", "Att", "Yds", "Y/A    TD", "Tgt", "   Rec", "Yds", "Y/R", "TD", "Ctch%", "Y/Tgt   TD", "Pts", "   Fmb", "FF   ", "FR", "Yds", "TD"]

    active_worksheet.append(labels)

    for x in row:
        items = x.findAll("td")
        counter = 0
        L = []

        append_row = [name]

        for y in items:
            if(counter == 0):
                if(y.string!="None"):
                    print(y.string)
                    append_row.append(y.text)
        active_worksheet.append(append_row)
        # print("---------------------------------------------------")
'''

# gets a list of the adjusted player names (first 4 letters of the last name)
def get_active_2018_urls(letter):

    url = "https://www.pro-football-reference.com/players/" + letter + "/"
    a = requests.get(url)           # get all the players with last name starting with letter
    soup = BeautifulSoup(a.text, 'lxml')
    content = soup.find("div", {"class":"section_content"})
    players_info = content.find_all("p")    # gets <p><b><a href="/players/Z/ZimmJu00.htm">Justin Zimmer</a> (DT)</b> 2018-2018</p>
                                        # or <p><a href="/players/W/WisnJe20.htm">Jerry Wisne</a> (T) 1999-2002</p> (if not marked as active with <b>)
    players_bold = []

    for p in players_info:
        p_split = p.text.split('-')
        last_year_played = int(p_split[-1])
        # Check if the player is marked as active and then validate it... might be able to get away with just the validation and not
        # checking for the "b" flag
        # p.b -- bolded name indicates active (<b><a href="/players/S/SaffRo20.htm">Rodger Saffold</a> (T)</b>)
        if p.b != None:
            if last_year_played == 2018:
                players_bold.append(p.b)
   
    # players_bold now holds a list of the following info: <b><a href="url to the player's page">name</a>(position)</b>
    # get the link for the webpage for the active players and add to gamelog_2018
    gamelog_2018 = []
    for p in players_bold:
        page = "https://www.pro-football-reference.com"
        page += p.a["href"]
        page += "/gamelog/2018"
        gamelog_2018.append(page)

    # maybe return a tuple with a list of names as well?
    return gamelog_2018



if __name__ == "__main__":

    wb = openpyxl.Workbook();
    dest_filename = 'data_2018.xlsx'
    active_worksheet = wb.active
    active_worksheet.title = "active players"

    # loop through each letter of the alphabet and get all the active players
    # (I'm trying to think of a more efficient way to do this bc this is pretty slow,
    # but... I don't think there is a better way)
    letters = list(string.ascii_uppercase)
    active_urls_2018 = []    # urls for 2018 game logs for every active player in the NFL
    
    for l in letters:
        active_urls_2018 += get_active_2018_urls(l)

    # print(get_active_2018_urls("A"))

    
    labels = ["Name", "Date", "G#", "Cmp", "Att", "Cmp%", "Yds", "TD", "Int", "Rate", "Sk", "Yds", "Y/A", "AY/A", "Att", "Yds", "Y/A    TD", "Tgt", "   Rec", "Yds", "Y/R", "TD", "Ctch%", "Y/Tgt   TD", "Pts", "   Fmb", "FF   ", "FR", "Yds", "TD"]
    # active_worksheet.append(labels)

    count = 0;

    
    for link in active_urls_2018:
        # active_worksheet.append(labels)
        # if count == 10:
        #     print("ending for testing purposes")
        #     break

        print("trying...", link)
        a = requests.get(link)
        soup = BeautifulSoup(a.text, 'lxml')

        try:
            tb = soup.find("tbody")
            row = tb.findAll("tr")
        except:
            # Gets an error here bc the player page lied to me and indicated a player as active by bolding the font
            # but the player's last active season is 2017, not 2018. We don't need this data, so we can skip it.
            print(count, "ERROR with", link)
            count +=1
            continue;

        for x in row:
            items = x.findAll("td")
            counter = 0
            # L = []

            append_row = [link]

            for y in items:
                if(counter == 0):
                    if(y.string!="None"):
                        # print(y.string)
                        append_row.append(y.text)

            active_worksheet.append(append_row)

    wb.save(filename = dest_filename)

    '''
     # getW = soup.find('span', {'itemprop': 'weight'})
        # weight = getW.text[0:3]
        # print(weight)
        # print("---------------------------------------------------")
    '''
