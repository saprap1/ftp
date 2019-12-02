#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Aug  7 10:57:56 2019

@authors: Jay Messina, Priya Sapra

@description: Scrape all the information for each player and store the data in spreadsheets
                
"""
import openpyxl
import os
from bs4 import BeautifulSoup
import requests
import pandas as pd

# Get the fantasy points for this player from the fantasy log online
def get_fant_points(link_half, year):

    # It broke at https://www.pro-football-reference.com/players/S/SpenDi00/fantasy/2017 because no fantasy log exists
    # Error: 
    '''
    Traceback (most recent call last):
      File "read_through_names.py", line 204, in <module>
        main()
      File "read_through_names.py", line 160, in main
        points = get_points("https://www.pro-football-reference.com/players/" + lastName[0] + "/" + lastName[:4] + firstName[:2] + "0" + str(version), year)
      File "read_through_names.py", line 90, in get_points
        for x in row:
    UnboundLocalError: local variable 'row' referenced before assignment
    '''

    link = link_half + "/fantasy/" + year
    a = requests.get(link)
    soup = BeautifulSoup(a.text, 'lxml')

    ret_points = []

    try:
        tb = soup.find("tbody")
        row = tb.findAll("tr")
    except:
        print("No fantasy log found for:", link)

    for x in row:
        items = x.findAll("td")
        
        # fantasy points are under the FantPt column, which is always the 3rd to last column
        try:
            pts = float(items[-3].text)
        except:
            # in case the cell is empty, let's just put 0 there... though this might skew training data
            pts = 0
        ret_points.append(pts)

    return ret_points

# write the data for the specified year to excelt sheets
def load_data(year):
    
    book2 = openpyxl.Workbook()
    book2 = openpyxl.load_workbook('all_players_2019.xlsx')
    sheet2 = book2['Sheet1']
    row_count2 = sheet2.max_row


    for i in range (945, row_count2+1):

        position = sheet2.cell(row=i, column=2).value
        s = sheet2.cell(row=i, column=1).value
        x = s.split(",")
        firstName = x[-1][1:]
        lastName = x[0]
        version = 0
        
        if firstName == "Derek" and lastName == "Carr":
            version = 2
        
        link = "https://www.pro-football-reference.com/players/" + lastName[0] + "/" + lastName[:4] + firstName[:2] + "0" + str(version) + "/gamelog/" + year + "/"

        
        ########################more debug stuff#######################
        #link = "https://www.pro-football-reference.com/players/G/GronRo00/gamelog/2018/"
        #firstName = "Rob"
        #lastName = "Gronkowski"

        # link = "https://www.pro-football-reference.com/players/B/BarkSa00/gamelog/2018"
        # firstName = "Saquon"
        # lastName = "Barkley"

        #link = "https://www.pro-football-reference.com/players/G/GronRo00/gamelog/2018/"
        #firstName = "Rob"
        #lastName = "Gronkowski"

        # link = "https://www.pro-football-reference.com/players/W/WalkHu00/gamelog/2018/"
        # firstName = "Delanie"
        # lastName = "Walker"

        # link = "https://www.pro-football-reference.com/players/E/EbroEr00/gamelog/2018/"
        # firstName = "Eric"
        # lastName = "Ebron"

        # link = "https://www.pro-football-reference.com/players/G/GrifGa00/gamelog/2018"
        # firstName = "Garret"
        # lastName = "Griffin"
        # position = "TE"
        ###############################################################

        # make soup request for data to formulate spreadsheet
        # open a spreadsheet, add fields, save it
        a = requests.get(link)
        soup = BeautifulSoup(a.text, 'lxml')


        # get ALL the fantasy points for this player in 1 soup request to the fantasy page
        # each entry is the fantasy points scored for that game
        fantasy_points = []
        fantasy_points = get_fant_points("https://www.pro-football-reference.com/players/" + lastName[0] + "/" + lastName[:4] + firstName[:2] + "0" + str(version), year)
        point_count = 0;    # pointer to current index of fantasy points

        try:
            tb = soup.find("tbody")
            row = tb.findAll("tr")

        except:
            # Here, need to handle 2 cases..
            #   - either player is new and doesn't have 2018/2017 data
            #   - or player has a different version # due to sharing a URL
            try:
                version = 2
                fantasy_points = get_fant_points("https://www.pro-football-reference.com/players/" + lastName[0] + "/" + lastName[:4] + firstName[:2] + "0" + str(version), year)
                tb = soup.find("tbody")
                row = tb.findAll("tr")
            except:
                try:
                    version = 3
                    fantasy_points = get_fant_points("https://www.pro-football-reference.com/players/" + lastName[0] + "/" + lastName[:4] + firstName[:2] + "0" + str(version), year)
                    tb = soup.find("tbody")
                    row = tb.findAll("tr")
                except:
                    print("skipping data collection for " + firstName + " " + lastName + "... (no data available or wrong version)")
                    continue
        
        # determine what features we want to focus on for each position
        features_wanted = []
        if position == "QB":
            features_wanted = ["pass_yds", "pass_td", "pass_int", "rush_yds", "rush_td"]
        elif position == "RB":
            features_wanted = ["rush_yds", "rush_td", "rec_yds", "rec_td", "fumbles"]
        elif position == "WR":
            features_wanted = ["rec_yds", "rec_td", "rush_yds", "rush_td", "fumbles"]
        elif position == "TE":
            features_wanted = ["rec_yds", "rec_td", "fumbles"]
        elif position == "K":
            features_wanted = ["xpm", "xpa", "fgm", "fga", "scoring"]
        else:
            # skip the player if it's not one that is specified (can adjust this later on...)
            print("skipping data collection for " + firstName + " " + lastName + "... (invalid position)")
            continue

        data = []
        
        # If the player has played fewer than 5 games, we'll skip them
        if (len(row) < 5):
            print("skipping data collection for " + firstName + " " + lastName + "... (played fewer than 5 games)")
            continue

        # get the player's height and weight and add it to data_append
        height = soup.find("span", itemprop="height")   # height in ft-in
        height = height.text
        height_list = height.split('-')
        height = float(height_list[0]) + float(height_list[1])/12.0     # convert the height to feet as a float

        weight = soup.find("span", itemprop="weight")   # weight in lbs
        weight = weight.text
        weight = float(weight[:-2])     # remove the last "lb" part of the string and convert to a float

        #print(len(row), "games recorded for", year, "for player:", firstName, lastName)
        print(len(row), "games recorded for", year)
        
        for x in row:
            data_append = []

            for f in features_wanted:
                try:
                    stat = x.find("td", {'data-stat': f})
                    data_append.append(float(stat.text))
                except:
                    print("data-stat", f, "not found. Skipping this feature for", firstName, lastName)
                    data_append.append(0.0) # need to change this at some point
                    continue
            
            # add height and weight as taken from before
            data_append.append(height)
            data_append.append(weight)

            # add the fantasy points to the data to be appended
            try:
                data_append.append(float(fantasy_points[point_count]))
            except:
                data_append.append(0.0)

            data.append(data_append)
            point_count += 1        

        # create dataframe from the list
        features_wanted.append("height")
        features_wanted.append("weight")
        features_wanted.append("fantasy points")

        # convert the two lists into a dataframe with features_wanted as the column names and data to the... acttual data
        df = pd.DataFrame.from_records(data, columns=features_wanted)

        dest_filename = firstName + "_" + lastName + "_" + position + "_" + year + ".xlsx"
        dest_path = year + "_data/" + dest_filename

        # Write the file out -- we should only get here if the player has played more than 5 games
        # and there was successful data collection for fantasy points and gamelog
        writer = pd.ExcelWriter(dest_path)
        df.to_excel(writer, 'Sheet1')
        writer.save()
    return
 
def main():
    load_data("2017")
    load_data("2018")
main()